from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from loguru import logger
from openai import OpenAI
from openpyxl import load_workbook


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
REQUEST_TIMEOUT_SECONDS = 5.0
MAX_TEXT_CHARS = 15000


def setup_logging() -> None:
    Path("logs").mkdir(parents=True, exist_ok=True)
    logger.remove()
    logger.add(
        "logs/scrape_{time:YYYY-MM-DD}.log",
        rotation="10 MB",
        retention="14 days",
        level="INFO",
        encoding="utf-8",
    )
    logger.add(lambda message: print(message, end=""), level="INFO")


def load_settings() -> tuple[str, str, str]:
    load_dotenv()
    api_key = os.getenv("AZURE_FOUNDRY_API_KEY", "").strip()
    endpoint = os.getenv("AZURE_FOUNDRY_ENDPOINT", "").strip()
    model = os.getenv("AZURE_FOUNDRY_MODEL", "").strip().strip('"')

    missing = [
        key
        for key, value in {
            "AZURE_FOUNDRY_API_KEY": api_key,
            "AZURE_FOUNDRY_ENDPOINT": endpoint,
            "AZURE_FOUNDRY_MODEL": model,
        }.items()
        if not value
    ]
    if missing:
        raise RuntimeError(f"Missing required environment variables: {', '.join(missing)}")

    return api_key, endpoint, model


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def find_header_row(sheet) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        index_by_name: dict[str, int] = {}
        for col_idx, cell_value in enumerate(row):
            header = normalize_header(cell_value)
            if header:
                index_by_name[header] = col_idx

        if "university" in index_by_name and "url" in index_by_name:
            return row_idx, index_by_name

    raise RuntimeError("Could not find header row with 'University' and 'url' columns.")


def read_targets(excel_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active

    header_row, index_by_name = find_header_row(sheet)
    university_idx = index_by_name["university"]
    url_idx = index_by_name["url"]
    restaurant_idx = index_by_name.get("restaurant")

    targets: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_url = row[url_idx] if url_idx < len(row) else None
        url = str(raw_url).strip() if raw_url else ""
        if not url:
            continue

        raw_university = row[university_idx] if university_idx < len(row) else None
        university = str(raw_university).strip() if raw_university else ""

        restaurant = ""
        if restaurant_idx is not None and restaurant_idx < len(row):
            raw_restaurant = row[restaurant_idx]
            restaurant = str(raw_restaurant).strip() if raw_restaurant else ""

        targets.append(
            {
                "university": university,
                "url": url,
                "restaurant_hint": restaurant,
            }
        )

    workbook.close()
    return targets


def fetch_readable_text(client: httpx.Client, url: str) -> str:
    response = client.get(url, timeout=REQUEST_TIMEOUT_SECONDS, follow_redirects=True)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg", "img"]):
        tag.decompose()

    text = soup.get_text(separator="\n")
    clean_lines = [line.strip() for line in text.splitlines() if line.strip()]
    clean_text = "\n".join(clean_lines)
    return clean_text[:MAX_TEXT_CHARS]


def parse_json_from_response(raw_text: str) -> dict[str, Any]:
    stripped = raw_text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?", "", stripped, flags=re.IGNORECASE).strip()
        stripped = re.sub(r"```$", "", stripped).strip()
    return json.loads(stripped)


def normalize_meal_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"breakfast", "lunch", "dinner"}:
        return text
    if text in {"morning", "brunch"}:
        return "breakfast"
    return "unknown"


def is_valid_meal_name(name: str) -> bool:
    cleaned = name.strip()
    if len(cleaned) < 2:
        return False

    lowered = cleaned.lower()
    invalid_tokens = {
        "menu",
        "today",
        "meal",
        "breakfast",
        "lunch",
        "dinner",
        "조식",
        "중식",
        "석식",
        "식단",
        "메뉴",
    }
    if lowered in invalid_tokens:
        return False

    if re.fullmatch(r"[0-9\-/:\s]+", cleaned):
        return False

    return True


def extract_menus_with_ai(
    ai_client: OpenAI,
    model: str,
    page_text: str,
    url: str,
    university: str,
    restaurant_hint: str,
) -> list[dict[str, Any]]:
    schema_instructions = {
        "items": [
            {
                "menu_date": "YYYY-MM-DD or empty string if unknown",
                "meal_type": "breakfast|lunch|dinner|unknown",
                "meal_name": "actual dish name only",
                "price_krw": "digits only string, or empty string if unknown",
                "restaurant_name": "restaurant or cafeteria name if known, else empty string",
                "is_confident_meal_name": "true if clearly a real meal/dish name, else false",
            }
        ]
    }

    prompt = (
        "Extract cafeteria menu items from the following webpage text and return strict JSON only.\n"
        "Rules:\n"
        "1) Return only valid JSON with this shape: "
        f"{json.dumps(schema_instructions, ensure_ascii=False)}\n"
        "2) Keep only real meal names. If uncertain, set is_confident_meal_name=false.\n"
        "3) meal_type must be breakfast, lunch, dinner, or unknown.\n"
        "4) price_krw must be digits as a string (e.g., 5000) or empty string if unknown.\n"
        "5) Do not invent data.\n"
        f"Context university: {university or 'unknown'}\n"
        f"Context restaurant hint: {restaurant_hint or 'none'}\n"
        f"Page URL: {url}\n\n"
        "Webpage text:\n"
        f"{page_text}"
    )

    response = ai_client.responses.create(model=model, input=prompt, timeout=5.0)
    output_text = response.output_text or "{}"
    payload = parse_json_from_response(output_text)
    items = payload.get("items", [])
    return items if isinstance(items, list) else []


def build_record(
    raw_item: dict[str, Any],
    scrape_date: str,
    url: str,
    university: str,
    restaurant_hint: str,
) -> dict[str, str] | None:
    meal_name = str(raw_item.get("meal_name", "")).strip()
    is_confident = bool(raw_item.get("is_confident_meal_name", False))

    if not meal_name or not is_confident or not is_valid_meal_name(meal_name):
        return None

    menu_date = str(raw_item.get("menu_date", "")).strip()
    if menu_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", menu_date):
        menu_date = ""

    price_krw = re.sub(r"\D", "", str(raw_item.get("price_krw", "")))

    restaurant_name = str(raw_item.get("restaurant_name", "")).strip()
    if not restaurant_name:
        restaurant_name = restaurant_hint

    return {
        "scrape_date": scrape_date,
        "url": url,
        "menu_date": menu_date,
        "meal_type": normalize_meal_type(raw_item.get("meal_type", "unknown")),
        "meal_name": meal_name,
        "price_krw": price_krw,
        "university": university,
        "restaurant_name": restaurant_name,
        "restaurant": restaurant_name,
    }


def main() -> None:
    setup_logging()
    scrape_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    scrape_day = datetime.now().strftime("%Y-%m-%d")

    results_dir = Path("results")
    results_dir.mkdir(parents=True, exist_ok=True)
    output_file = results_dir / f"menus-{scrape_day}.jsonl"

    logger.info("Starting scrape process")
    api_key, endpoint, model = load_settings()

    excel_path = Path("campus_restaurant_websites.xlsx")
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    targets = read_targets(excel_path)
    logger.info(f"Loaded {len(targets)} target URLs from Excel")

    ai_client = OpenAI(api_key=api_key, base_url=endpoint)

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    }

    saved_count = 0
    with httpx.Client(headers=headers, timeout=REQUEST_TIMEOUT_SECONDS) as http_client:
        with output_file.open("a", encoding="utf-8") as sink:
            for idx, target in enumerate(targets, start=1):
                url = target["url"]
                university = target["university"]
                restaurant_hint = target["restaurant_hint"]

                logger.info(f"[{idx}/{len(targets)}] Processing {url}")

                try:
                    page_text = fetch_readable_text(http_client, url)
                    if not page_text:
                        logger.warning(f"Empty page text, skipped: {url}")
                        continue

                    raw_items = extract_menus_with_ai(
                        ai_client=ai_client,
                        model=model,
                        page_text=page_text,
                        url=url,
                        university=university,
                        restaurant_hint=restaurant_hint,
                    )

                    if not raw_items:
                        logger.info(f"No menu items extracted: {url}")
                        continue

                    site_saved = 0
                    for raw_item in raw_items:
                        if not isinstance(raw_item, dict):
                            continue
                        record = build_record(
                            raw_item=raw_item,
                            scrape_date=scrape_timestamp,
                            url=url,
                            university=university,
                            restaurant_hint=restaurant_hint,
                        )
                        if not record:
                            continue

                        sink.write(json.dumps(record, ensure_ascii=False) + "\n")
                        sink.flush()
                        saved_count += 1
                        site_saved += 1
                        logger.info(
                            f"Saved item #{saved_count}: {record['meal_name']} ({record['meal_type']})"
                        )

                    logger.info(f"Completed {url} with {site_saved} saved items")

                except httpx.TimeoutException:
                    logger.error(f"Timeout after {REQUEST_TIMEOUT_SECONDS}s, skipped: {url}")
                    continue
                except httpx.HTTPError as exc:
                    logger.error(f"HTTP error for {url}: {exc}")
                    continue
                except Exception as exc:
                    logger.exception(f"Failed processing {url}: {exc}")
                    continue

    logger.info(f"Finished. Saved {saved_count} items to {output_file}")


if __name__ == "__main__":
    main()